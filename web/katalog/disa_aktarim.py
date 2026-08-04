"""CSV ve XLSX dışa aktarım yanıtlarını güvenli ve bellek dostu üretir.

View'lar bu modüle yalnız dosya bilgilerini, sütun tanımlarını ve satır üreten bir
iterable verir. Sorgunun kurulması ve kullanıcıya gösterilecek değerlerin seçilmesi
view'ın sorumluluğundadır; böylece bu yardımcı filtre veya iş kuralı kopyalamaz.
"""

import csv
import re
from datetime import date, datetime, time
from tempfile import SpooledTemporaryFile

from django.http import FileResponse, Http404, StreamingHttpResponse
from django.utils import timezone
from django.utils.http import content_disposition_header


XLSX_ICERIK_TURU = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# Küçük dosyalar RAM'de kalır; bu eşiği aşan XLSX dosyaları otomatik olarak geçici
# diske taşınır. openpyxl'in write-only modu da çalışma sayfası satırlarını bellekte
# biriktirmez.
XLSX_BELLEK_ESIGI = 8 * 1024 * 1024

FORMUL_BASLANGICLARI = frozenset(('=', '+', '-', '@'))
GECERSIZ_SAYFA_KARAKTERLERI = re.compile(r'[\\/*?:\[\]]')


class _CsvYankisi:
    """csv.writer'ın ürettiği metni doğrudan geri veren dosya benzeri nesne."""

    @staticmethod
    def write(deger):
        return deger


def _metni_formulden_koru(deger):
    """Excel'in formül sayabileceği kullanıcı metnini düz metne dönüştür."""
    if not isinstance(deger, str):
        return deger

    # Baştaki boşluklar bazı tablo uygulamalarında ayrıştırma sırasında atılabilir;
    # kontrolü ilk görünür karaktere göre yapmak bu dolanma yolunu da kapatır.
    kirpilmis = deger.lstrip()
    if kirpilmis and kirpilmis[0] in FORMUL_BASLANGICLARI:
        return f"'{deger}"
    return deger


def _yerel_naive_tarih(deger):
    """Aware datetime'ı proje saat dilimine çevirip Excel için naive döndür."""
    if isinstance(deger, datetime) and timezone.is_aware(deger):
        deger = timezone.localtime(deger, timezone.get_default_timezone())
        return deger.replace(tzinfo=None)
    return deger


def _csv_degeri(deger):
    """Bir Python değerini CSV'de güvenli ve kararlı gösterime çevir."""
    deger = _yerel_naive_tarih(deger)
    if isinstance(deger, datetime):
        return deger.isoformat(sep=' ', timespec='seconds')
    if isinstance(deger, time):
        return deger.isoformat(timespec='seconds')
    if isinstance(deger, date):
        return deger.isoformat()
    if deger is None:
        return ''
    return _metni_formulden_koru(deger)


def _sutunlari_dogrula(sutunlar):
    """Sütun iterable'ını bir kez tüketip başlık ve genişlikleri doğrula."""
    sonuc = []
    for sutun in sutunlar:
        try:
            baslik, genislik = sutun
        except (TypeError, ValueError) as exc:
            raise ValueError('Her sütun (başlık, genişlik) ikilisi olmalıdır.') from exc

        baslik = str(baslik)
        try:
            genislik = float(genislik)
        except (TypeError, ValueError) as exc:
            raise ValueError(f'{baslik!r} sütununun genişliği sayısal olmalıdır.') from exc
        if genislik <= 0:
            raise ValueError(f'{baslik!r} sütununun genişliği sıfırdan büyük olmalıdır.')
        sonuc.append((baslik, genislik))

    if not sonuc:
        raise ValueError('Dışa aktarım için en az bir sütun tanımlanmalıdır.')
    return tuple(sonuc)


def _satiri_dogrula(satir, sutun_sayisi):
    """Tek satırı hücre dizisine çevirip sütun sayısının şaşmasını engelle."""
    try:
        degerler = tuple(satir)
    except TypeError as exc:
        raise ValueError('Dışa aktarım satırları iterable olmalıdır.') from exc
    if len(degerler) != sutun_sayisi:
        raise ValueError(
            'Dışa aktarım satırındaki değer sayısı sütun sayısıyla eşleşmiyor: '
            f'{len(degerler)} != {sutun_sayisi}.'
        )
    return degerler


def _dosya_adi(dosya_koku, uzanti):
    """Content-Disposition içinde güvenle kullanılacak bir dosya adı üret."""
    # İsim diske yazılmıyor; yine de satır sonlarını ve yol parçalarını başlığa
    # taşımamak hem güvenli hem de tarayıcılar arasında daha öngörülebilir.
    temiz_kok = str(dosya_koku).replace('\\', '-').replace('/', '-').strip()
    temiz_kok = ' '.join(temiz_kok.split())
    temiz_kok = temiz_kok.strip('.') or 'disa-aktarim'
    return f'{temiz_kok}.{uzanti}'


def _calisma_sayfasi_adi(ad):
    """Excel'in 31 karakter ve özel karakter sınırlarına uygun sayfa adı üret."""
    temiz_ad = GECERSIZ_SAYFA_KARAKTERLERI.sub('-', str(ad)).strip().strip("'")
    return temiz_ad[:31] or 'Dışa Aktarım'


def _csv_yaniti(dosya_koku, kapsam, sutunlar, satirlar):
    """UTF-8 BOM'lu ve noktalı virgülle ayrılmış akış yanıtı üret."""
    sutun_sayisi = len(sutunlar)

    def icerik():
        # BOM ayrı bir ilk parça olarak yazılır; StreamingHttpResponse bunu UTF-8
        # baytlarına çevirir ve Türkçe Excel kurulumları kodlamayı doğru tanır.
        yield '\ufeff'
        yazici = csv.writer(_CsvYankisi(), delimiter=';', lineterminator='\r\n')

        kapsam_satiri = [f'Kapsam: {_metni_formulden_koru(str(kapsam))}']
        kapsam_satiri.extend('' for _ in range(sutun_sayisi - 1))
        yield yazici.writerow(kapsam_satiri)
        yield yazici.writerow([baslik for baslik, _ in sutunlar])

        for satir in satirlar:
            degerler = _satiri_dogrula(satir, sutun_sayisi)
            yield yazici.writerow([_csv_degeri(deger) for deger in degerler])

    dosya_adi = _dosya_adi(dosya_koku, 'csv')
    yanit = StreamingHttpResponse(icerik(), content_type='text/csv; charset=utf-8')
    yanit.headers['Content-Disposition'] = content_disposition_header(True, dosya_adi)
    return yanit


def _xlsx_hucresi(calisma_sayfasi, deger):
    """Gerçek veri tipini ve güvenli metin tipini koruyan write-only hücre üret."""
    from openpyxl.cell import WriteOnlyCell

    deger = _yerel_naive_tarih(deger)
    hucre = WriteOnlyCell(calisma_sayfasi, value=deger)

    # openpyxl '=' ile başlayan metni kendiliğinden formül yapar. Hücreyi açıkça
    # metin tipine sabitlemek içeriği değiştirmeden formül çalıştırılmasını engeller;
    # başında sıfır bulunan stok kodları da böylece aynen korunur.
    if isinstance(deger, str):
        hucre.data_type = 's'
    elif isinstance(deger, datetime):
        hucre.number_format = 'yyyy-mm-dd hh:mm:ss'
    elif isinstance(deger, date):
        hucre.number_format = 'yyyy-mm-dd'
    elif isinstance(deger, time):
        hucre.number_format = 'hh:mm:ss'
    return hucre


def _xlsx_yaniti(dosya_koku, calisma_sayfasi, kapsam, sutunlar, satirlar):
    """Write-only çalışma kitabını sınırlı bellekle FileResponse'a dönüştür."""
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    kitap = Workbook(write_only=True)
    sayfa = kitap.create_sheet(_calisma_sayfasi_adi(calisma_sayfasi))
    son_sutun = get_column_letter(len(sutunlar))

    # Write-only çalışma sayfasında görünüm ayarları ilk hücrelerden önce yapılmalı.
    sayfa.freeze_panes = 'A3'
    for sira, (_, genislik) in enumerate(sutunlar, start=1):
        sayfa.column_dimensions[get_column_letter(sira)].width = genislik

    kapsam_hucreleri = []
    for sira in range(len(sutunlar)):
        deger = f'Kapsam: {kapsam}' if sira == 0 else ''
        hucre = WriteOnlyCell(sayfa, value=deger)
        hucre.data_type = 's'
        hucre.font = Font(bold=True, italic=True, color='374151')
        kapsam_hucreleri.append(hucre)
    sayfa.append(kapsam_hucreleri)

    baslik_hucreleri = []
    for baslik, _ in sutunlar:
        hucre = WriteOnlyCell(sayfa, value=baslik)
        hucre.data_type = 's'
        hucre.font = Font(bold=True, color='FFFFFF')
        hucre.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        hucre.alignment = Alignment(vertical='center')
        baslik_hucreleri.append(hucre)
    sayfa.append(baslik_hucreleri)

    veri_satiri_sayisi = 0
    for satir in satirlar:
        degerler = _satiri_dogrula(satir, len(sutunlar))
        sayfa.append([_xlsx_hucresi(sayfa, deger) for deger in degerler])
        veri_satiri_sayisi += 1

    # Sonuç boşken de başlık satırı filtre aralığının içinde kalır.
    son_satir = 2 + veri_satiri_sayisi
    sayfa.auto_filter.ref = f'A2:{son_sutun}{son_satir}'

    gecici_dosya = SpooledTemporaryFile(max_size=XLSX_BELLEK_ESIGI, mode='w+b')
    try:
        kitap.save(gecici_dosya)
        gecici_dosya.seek(0)
    except Exception:
        gecici_dosya.close()
        raise

    return FileResponse(
        gecici_dosya,
        as_attachment=True,
        filename=_dosya_adi(dosya_koku, 'xlsx'),
        content_type=XLSX_ICERIK_TURU,
    )


def dosya_yaniti(dosya_turu, dosya_koku, calisma_sayfasi, kapsam, sutunlar, satirlar):
    """İstenen biçimde dışa aktarım yanıtı üret.

    ``sutunlar`` her biri ``(başlık, Excel genişliği)`` olan ikililerden;
    ``satirlar`` ise aynı sayıda hücre içeren iterable satırlardan oluşur. Satırlar
    listeye çevrilmez: CSV doğrudan akar, XLSX ise write-only çalışma kitabına yazılır.
    """
    tur = str(dosya_turu).strip().lower()
    if tur not in {'csv', 'xlsx'}:
        raise Http404('Desteklenmeyen dışa aktarım biçimi.')

    dogrulanmis_sutunlar = _sutunlari_dogrula(sutunlar)
    if tur == 'csv':
        return _csv_yaniti(dosya_koku, kapsam, dogrulanmis_sutunlar, satirlar)
    return _xlsx_yaniti(
        dosya_koku,
        calisma_sayfasi,
        kapsam,
        dogrulanmis_sutunlar,
        satirlar,
    )
