"""Fotoğrafı olan ürünleri, küçültülmüş (thumbnail) gömülü görsellerle tek bir
Excel dosyasında birleştirir — urun_listesi.xlsx'in (stok kodu + görsel yan
yana) temiz/güncel bir hâli, ama bu sefer veritabanı doğrulamasından geçmiş
2.974 ürünün fotoğrafı olan 1.780 tanesiyle sınırlı (2026-07-28).

Orijinal urun_listesi.xlsx tam çözünürlükte gömülü görseller yüzünden
~195 MB'tı; burada thumbnail kullanılarak dosya boyutu birkaç MB'a indirilir.
"""

from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
import psycopg2
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


BASE_DIR = Path(__file__).resolve().parents[2]
GORSEL_KLASORU = BASE_DIR / "images" / "final" / "products"
CIKIS_DOSYASI = BASE_DIR / "data" / "processed" / "urun_katalogu_gorselli.xlsx"

DB_NAME = "depo_sistemi"
DB_USER = "depo_admin"
DB_PASSWORD = "supergizlisifre"
DB_HOST = "localhost"
DB_PORT = 5433

THUMBNAIL_PX = 120
SATIR_YUKSEKLIGI_PT = 92
GORSEL_SUTUN_GENISLIGI = 18


def veriyi_oku() -> pd.DataFrame:
    conn = psycopg2.connect(
        dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD, host=DB_HOST, port=DB_PORT
    )
    try:
        df = pd.read_sql(
            """
            SELECT
                u.stok_kodu, k.kategori_adi, u.urun_tipi, u.olcu_mm,
                u.boy_ligne, u.aciklama, g.dosya_adi
            FROM urunler u
            JOIN urun_gorselleri g
                ON g.stok_kodu = u.stok_kodu AND g.ana_gorsel_mi = TRUE
            LEFT JOIN kategoriler k ON k.kategori_id = u.kategori_id
            ORDER BY u.stok_kodu;
            """,
            conn,
        )
    finally:
        conn.close()
    return df


def thumbnail_olustur(dosya_yolu: Path) -> BytesIO | None:
    try:
        with PILImage.open(dosya_yolu) as img:
            img = img.convert("RGBA") if img.mode in ("P", "LA") else img.convert("RGB")
            img.thumbnail((THUMBNAIL_PX, THUMBNAIL_PX))
            buf = BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            return buf
    except Exception as exc:
        print(f"⚠️  Görsel işlenemedi ({dosya_yolu.name}): {exc}")
        return None


def main() -> None:
    df = veriyi_oku()
    print(f"📦 Fotoğrafı olan ürün sayısı: {len(df)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urun_Katalogu"

    basliklar = [
        "GÖRSEL", "ÜRÜN STOK KODU", "ÜRÜN KATEGORİ GRUBU",
        "ÖLÇÜ (mm)", "BOY (ligne)", "AÇIKLAMA", "GÖRSEL DOSYA ADI",
    ]
    ws.append(basliklar)
    ws.column_dimensions[get_column_letter(1)].width = GORSEL_SUTUN_GENISLIGI
    for idx in range(2, len(basliklar) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20

    eksik_dosya = 0
    islenen = 0

    for i, row in df.iterrows():
        excel_satiri = i + 2
        ws.cell(row=excel_satiri, column=2, value=row["stok_kodu"])
        ws.cell(row=excel_satiri, column=3, value=row["kategori_adi"])
        ws.cell(row=excel_satiri, column=4, value=row["olcu_mm"])
        ws.cell(row=excel_satiri, column=5, value=row["boy_ligne"])
        ws.cell(row=excel_satiri, column=6, value=row["aciklama"])
        ws.cell(row=excel_satiri, column=7, value=row["dosya_adi"])
        ws.row_dimensions[excel_satiri].height = SATIR_YUKSEKLIGI_PT

        dosya_yolu = GORSEL_KLASORU / row["dosya_adi"]
        if not dosya_yolu.exists():
            eksik_dosya += 1
            continue

        buf = thumbnail_olustur(dosya_yolu)
        if buf is None:
            continue

        xl_img = XLImage(buf)
        ws.add_image(xl_img, f"A{excel_satiri}")
        islenen += 1

        if islenen % 200 == 0:
            print(f"   ... {islenen}/{len(df)} görsel işlendi")

    wb.save(CIKIS_DOSYASI)

    print(f"✅ Katalog oluşturuldu: {CIKIS_DOSYASI}")
    print(f"🖼️ Gömülen görsel: {islenen}")
    if eksik_dosya:
        print(f"⚠️ Bulunamayan görsel dosyası: {eksik_dosya}")
    print(f"📏 Dosya boyutu: {CIKIS_DOSYASI.stat().st_size / 1e6:.1f} MB")


if __name__ == "__main__":
    main()
